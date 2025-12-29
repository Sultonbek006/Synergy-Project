"""
Excel Parser Service
Handles parsing of Excel files with Russian headers and inserting into database
"""
import re
from io import BytesIO
from typing import List, Dict, Any

from openpyxl import load_workbook
from sqlalchemy.orm import Session

from .models import MasterPlan


# Region normalization mapping (Cyrillic/Latin -> Standardized Latin Uppercase)
# Region normalization mapping (Cyrillic/Latin -> Standardized Latin Uppercase)
# Region normalization mapping (Cyrillic/Latin -> Standardized Latin Uppercase)
REGION_MAP = {
    # === TASHKENT OBL (Includes Sirdaryo & Exact Excel Matches) ===
    'ТАШКЕНТ (ОБЛ)': 'TOSHKENT OBL',   # Exact match from Excel
    'ТАШКЕНТ(ОБЛ)': 'TOSHKENT OBL',    # Variant without space
    'TASHKENT (OBL)': 'TOSHKENT OBL',
    'СЫРДАРЬЯ': 'TOSHKENT OBL',
    'СИРДАРЁ': 'TOSHKENT OBL',
    'SIRDARYO': 'TOSHKENT OBL',
    'SIR': 'TOSHKENT OBL',
    'SYRDARYA': 'TOSHKENT OBL',
    'GULISTON': 'TOSHKENT OBL',
    'YANGIYER': 'TOSHKENT OBL',
    'YANGIYO\'L': 'TOSHKENT OBL',      # Explicit city check
    'YANGIYUL': 'TOSHKENT OBL',
    'ANGREN': 'TOSHKENT OBL',
    'CHIRCHIQ': 'TOSHKENT OBL',
    'OLMALIQ': 'TOSHKENT OBL',
    'BEKOBOD': 'TOSHKENT OBL',
    'BUKA': 'TOSHKENT OBL',
    'BOKA': 'TOSHKENT OBL',
    
    'ТАШКЕНТ ОБЛ': 'TOSHKENT OBL',
    'ТАШ ОБЛ': 'TOSHKENT OBL',
    'ТАШ.ОБЛ': 'TOSHKENT OBL',
    'TOSHKENT OBL': 'TOSHKENT OBL',
    'TOSH OBL': 'TOSHKENT OBL',
    'TOSH.OBL': 'TOSHKENT OBL',
    'TOSHKENT VIL': 'TOSHKENT OBL',
    'TOSH VIL': 'TOSHKENT OBL',
    'T.VIL': 'TOSHKENT OBL',
    'VILOYAT': 'TOSHKENT OBL', 
    
    # === TASHKENT OBSHIY (General) ===
    'ТАШКЕНТ (ОБЩ)': 'TOSHKENT OBSH',  # Exact match from Excel
    'ТАШКЕНТ(ОБЩ)': 'TOSHKENT OBSH',
    'TASHKENT (OBSH)': 'TOSHKENT OBSH',
    'ОБЩИЙ': 'TOSHKENT OBSH',
    'ТАШКЕНТ ОБЩ': 'TOSHKENT OBSH',
    'TOSHKENT OBSH': 'TOSHKENT OBSH',
    'OBSH': 'TOSHKENT OBSH',
    'OBS': 'TOSHKENT OBSH',
    'UMUMIY': 'TOSHKENT OBSH',
    'GENERAL': 'TOSHKENT OBSH',
    
    # === TASHKENT CITY (Must be checked AFTER Obl/Obsh) ===
    'ТАШКЕНТ': 'TOSHKENT CITY',
    'Г.ТАШКЕНТ': 'TOSHKENT CITY',
    'Т.Г': 'TOSHKENT CITY',
    'TOSHKENT': 'TOSHKENT CITY',
    'TOSH': 'TOSHKENT CITY',
    'TASHKENT': 'TOSHKENT CITY',
    'TASH': 'TOSHKENT CITY',

    # === STANDARD REGIONS ===
    'СУРХАНДАРЬЯ': 'SURXANDARYO',
    'КАШКАДАРЬЯ': 'QASHQADARYO',
    'САМАРКАНД': 'SAMARQAND',
    'БУХАРА': 'BUXORO',
    'НАМАНГАН': 'NAMANGAN',
    'АНДИЖАН': 'ANDIJON',
    'ФЕРГАНА': "FARG'ONA",
    'ДЖИЗАК': 'JIZZAX',
    'НАВОИ': 'NAVOIY',
    'ХОРЕЗМ': 'XORAZM',
    'НУКУС': 'NUKUS',
    'KARAKALPAKSTAN': 'NUKUS',
    'QORAQALPOGISTON': 'NUKUS',

    # Latin matches
    'SURXANDARYO': 'SURXANDARYO',
    'QASHQADARYO': 'QASHQADARYO',
    'SAMARQAND': 'SAMARQAND',
    'BUXORO': 'BUXORO',
    'NAMANGAN': 'NAMANGAN',
    'ANDIJON': 'ANDIJON',
    "FARG'ONA": "FARG'ONA",
    'JIZZAX': 'JIZZAX',
    'NAVOIY': 'NAVOIY',
    'XORAZM': 'XORAZM',
    'NUKUS': 'NUKUS',
    
    # Precise Short Matches
    'SUR': 'SURXANDARYO',
    'QASH': 'QASHQADARYO',
    'SAM': 'SAMARQAND',
    'BUX': 'BUXORO',
    'NAM': 'NAMANGAN',
    'AND': 'ANDIJON',
    'FERG': "FARG'ONA",
    'FARG': "FARG'ONA",
    'JIZ': 'JIZZAX',
    'NAV': 'NAVOIY',
    'XOR': 'XORAZM',
    'NUK': 'NUKUS',
}


def normalize_region(region_str: str) -> str:
    """Normalize region name to standardized uppercase Latin"""
    if not region_str:
        return 'UNKNOWN'
    
    normalized = region_str.strip().upper()
    
    # 1. Direct match (Fast path)
    if normalized in REGION_MAP:
        return REGION_MAP[normalized]
    
    # 2. Sort keys by length descending to ensure "TOSH OBL" is matched before "TOSH"
    sorted_keys = sorted(REGION_MAP.keys(), key=len, reverse=True)
    
    for key in sorted_keys:
        if key in normalized:
            return REGION_MAP[key]
    
    # Return as-is if no match
    return normalized



def clean_phone(phone_str: str) -> str:
    """Extract only digits from phone number"""
    if not phone_str:
        return ''
    return re.sub(r'[^0-9]', '', str(phone_str))


def clean_amount(amount_value: Any) -> int:
    """Clean and convert amount to integer"""
    if amount_value is None:
        return 0
    
    if isinstance(amount_value, (int, float)):
        return int(amount_value)
    
    # Remove spaces, commas, text
    cleaned = re.sub(r'[^0-9.-]', '', str(amount_value))
    if not cleaned:
        return 0
    
    try:
        return int(float(cleaned))
    except ValueError:
        return 0


# Field mapping (Internal Field Name -> List of possible Excel Headers)
HEADER_MAP = {
    'doctor_name': ['фио', 'doctor', 'name', 'full name', 'имя', 'fullname'],
    'region': ['регион', 'region', 'viloyat'],
    'district': ['район', 'district', 'tuman', 'city'],
    'target_amount': ['сумм', 'сумма', 'target', 'amount', 'plan', 'actual amount', 'total', 'summ', 'план'],
    'planned_type': ['форма', 'type', 'form', 'payment type', 'to\'lov turi'],
    'card_number': ['номер карты', 'card', 'card number', 'karta'],
    'workplace': ['место работы', 'workplace', 'work', 'joy'],
    'specialty': ['специальность', 'specialty', 'spec', 'kasb'],
    'phone': ['номер телефо', 'phone', 'tel', 'number', 'mobile', 'телефон'],
    'group_name': ['групп', 'group', 'guruh', 'sinf', 'toifa'],
    'manager_name': ['мп', 'manager', 'rm', 'regional manager', 'boshqaruvchi'],
}

# Column Index Fallback (If headers not found, use these columns - 1-based)
# Updated for new 10-column format: A=Doc, B=Reg, C=Dist, D=Amt, E=Type, F=Workplace, G=Specialty, H=Phone, I=Group, J=RM
COLUMN_MAP_FALLBACK = {
    'doctor_name': 1,    # Column A - ФИО
    'region': 2,         # Column B - Регион
    'district': 3,       # Column C - Район
    'target_amount': 4,  # Column D - Сумм
    'planned_type': 5,   # Column E - Форма
    'workplace': 6,      # Column F - Место работы (NEW)
    'specialty': 7,      # Column G - Специальность (NEW)
    'phone': 8,          # Column H - Номер телефона (moved from F)
    'group_name': 9,     # Column I - Группа (moved from G)
    'manager_name': 10,  # Column J - МП/RM (moved from H)
    # Optional - not in standard format
    'card_number': 0,
}

def find_header_row(ws) -> tuple:
    """Find the header row and create column mapping"""
    
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=True), 1):
        col_mapping = {}
        non_empty_count = 0
        
        for col_idx, cell_value in enumerate(row, 1):
            if cell_value is None:
                continue
            non_empty_count += 1
            cell_lower = str(cell_value).lower().strip()
            
            for field, patterns in HEADER_MAP.items():
                for pattern in patterns:
                    if pattern in cell_lower:
                        col_mapping[field] = col_idx
                        break
        
        # If we found at least 3 column keywords, this is likely the header
        if len(col_mapping) >= 3:
            return row_idx, col_mapping
    
    # Check first row to detect data format
    first_row = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0] if ws.max_row > 0 else []
    non_empty_cols = sum(1 for cell in first_row if cell is not None)
    
    # If first row looks like data (has a number in column D which would be target_amount)
    # and has 4-12 columns, use 10-column fallback
    if non_empty_cols >= 4 and non_empty_cols <= 12:
        # Check if column D looks like an amount (numeric)
        if len(first_row) >= 4:
            col_d = first_row[3] if len(first_row) > 3 else None
            if col_d is not None:
                try:
                    # If it's numeric, this is likely data without headers
                    float(str(col_d).replace(',', '').replace(' ', ''))
                    return 0, COLUMN_MAP_FALLBACK
                except (ValueError, TypeError):
                    pass
    
    # Default fallback: start from row 1 with 10-column layout
    return 0, COLUMN_MAP_FALLBACK



def process_excel_file(
    file_content: bytes,
    company_name: str,
    db: Session,
    month: int = 12  # Default to December
) -> Dict[str, Any]:
    """
    Parse Excel file and insert rows into master_plan table
    
    Returns:
        Dict with 'success', 'inserted_count', 'errors'
    """
    try:
        # Load workbook from bytes
        wb = load_workbook(filename=BytesIO(file_content), data_only=True)
        ws = wb.active
        
        # Find headers
        header_row, col_map = find_header_row(ws)
        
        inserted_count = 0
        errors = []
        
        # Iterate through data rows (skip header)
        for row_idx, row in enumerate(ws.iter_rows(min_row=header_row + 1, values_only=True), header_row + 1):
            try:
                # Extract values using column mapping
                def get_val(field: str) -> Any:
                    col_idx = col_map.get(field)
                    if col_idx and col_idx <= len(row):
                        return row[col_idx - 1]
                    return None
                
                doctor_name = get_val('doctor_name')
                target_amount = clean_amount(get_val('target_amount'))
                
                # Skip empty rows
                if not doctor_name and target_amount == 0:
                    continue
                
                # Skip header-like rows
                if doctor_name and str(doctor_name).lower().strip() in ['фио', 'name', 'doctor']:
                    continue
                
                # Create record
                plan_item = MasterPlan(
                    company=company_name,
                    doctor_name=str(doctor_name).strip() if doctor_name else 'Unknown',
                    region=normalize_region(str(get_val('region') or '')),
                    district=str(get_val('district') or '').strip(),
                    target_amount=target_amount,
                    planned_type=str(get_val('planned_type') or 'Cash').strip(),
                    card_number=str(get_val('card_number') or '').strip(),
                    workplace=str(get_val('workplace') or '').strip(),
                    specialty=str(get_val('specialty') or '').strip(),
                    phone=clean_phone(get_val('phone')),
                    group_name=str(get_val('group_name') or 'UNASSIGNED').strip().upper(),
                    manager_name=str(get_val('manager_name') or '').strip(),
                    month=month,
                    status='Pending'
                )
                
                db.add(plan_item)
                inserted_count += 1
                
            except Exception as e:
                errors.append(f"Row {row_idx}: {str(e)}")
        
        db.commit()
        
        return {
            'success': True,
            'inserted_count': inserted_count,
            'errors': errors
        }
        
    except Exception as e:
        db.rollback()
        return {
            'success': False,
            'inserted_count': 0,
            'errors': [str(e)]
        }
